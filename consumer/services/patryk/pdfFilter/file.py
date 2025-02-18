from config.app_config import FILE, TMP
import os
import zipfile
from PIL import Image
from io import BytesIO
from openpyxl import load_workbook
import re
import shutil
from config.app_config import START_LITERAL_COLUMN_FILE_XLSX


def extract_images_from_xlsx() -> tuple[list[dict], str, bool]:
    try:
        os.makedirs(TMP, exist_ok=True)
        data = []

        file_path, is_valid = file_name_find()
        if not is_valid:
            return data, file_path, False

        file_path = FILE / file_path

        with zipfile.ZipFile(file_path, 'r') as xlsx_zip:
            media_files = sorted(
                [f for f in xlsx_zip.namelist() if f.startswith('xl/media/')]
            )

            def extract_number_from_filename(filename):
                match = re.search(r'(\d+)', filename)
                return int(match.group(1)) if match else None

            for file_name in media_files:
                try:
                    image_data = xlsx_zip.read(file_name)
                    image = Image.open(BytesIO(image_data))

                    image_number = extract_number_from_filename(os.path.basename(file_name))

                    if image_number is None:
                        continue

                    image_name = f"image{image_number}.png"
                    save_path = os.path.join(TMP, image_name)

                    image.save(save_path, format='PNG')

                    new_image_obj = {
                        "index": image_number,
                        "name_file": image_name,
                        "url_image": save_path,
                    }
                    data.append(new_image_obj)
                except Exception as e:
                    print(f"Image processing error {file_name}: {e}")

        return data, "", True

    except Exception as e:
        return [], str(e), False


def file_name_find(file_type: str = 'xlsx') -> tuple[str, bool]:
    try:
        if not os.path.isdir(FILE):
            return f"Path {FILE} is not a folder or does not exist", False

        if file_type == 'xlsx':
            for file in os.listdir(FILE):
                if file.endswith(".xlsx"):
                    return file, True
        else:
            for file in os.listdir(FILE):
                if file.endswith(".xls"):
                    return file, True

        return "No .xlsx or .xls files in folder", False
    except Exception as e:
        return str(e), False


def open_file_xlsx() -> tuple[list[dict], str, bool]:
    try:
        name_file, success = file_name_find()
        path = os.path.join(FILE, name_file)

        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active

        headers = [cell.value for cell in sheet[START_LITERAL_COLUMN_FILE_XLSX - 1]]
        headers = [header.strip() for header in headers if header and isinstance(header, str)]

        data = []
        index = 1

        for row_idx, row in enumerate(sheet.iter_rows(min_row=10, values_only=True),
                                      start=START_LITERAL_COLUMN_FILE_XLSX):
            if not any(row):
                continue

            row_data = {headers[i]: row[i] for i in range(len(headers)) if i < len(row) and headers[i]}

            if row_data.get('Lp') == 'Lp':
                continue

            new_object = {
                'index': index,
                'column_excel': row_idx,
                'lp': row_data.get('Lp', ''),
                'name': row_data.get('Nazwa towaru', ''),
                'quantity': row_data.get('Ilość', ''),
                'ean': row_data.get('EAN', ''),
                'location': row_data.get('Lokalizacja', ''),
            }

            data.append(new_object)
            index += 1

        return data, "", True

    except Exception as e:
        return [], str(e), False


def clear_tmp_files(file_paths: list[str]) -> tuple[str, bool]:
    try:
        for file_path in file_paths:
            if os.path.exists(file_path):
                os.remove(file_path)
        return "", True
    except Exception as e:
        return str(e), False


def clear_catalog(path_catalog) -> tuple[str, bool]:
    try:
        path_file_catalog = path_catalog
        if not path_file_catalog.exists():
            return "The directory does not exist.", False

        for item in path_file_catalog.iterdir():
            if item.is_file() and not item.suffix == ".py":
                item.unlink()

            elif item.is_dir():
                shutil.rmtree(item)

        return "The directory has been cleared.", True

    except Exception as e:
        return str(e), False